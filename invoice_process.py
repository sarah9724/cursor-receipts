import os
import shutil
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import TYPE_NUMERIC, TYPE_STRING
from datetime import datetime
import re
import tempfile
from werkzeug.utils import secure_filename
import io

# 添加文件大小检查
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

def extract_invoice_data(pdf):
    """提取PDF中的发票数据"""
    first_page_text = pdf.pages[0].extract_text()
    full_text = ""
    top_right_text = pdf.pages[0].crop((pdf.pages[0].width * 0.5, 0, pdf.pages[0].width, pdf.pages[0].height * 0.2)).extract_text()
    
    for page in pdf.pages:
        full_text += page.extract_text() + "\n"
 
    return first_page_text, full_text, top_right_text

def parse_data(first_page_text, full_text, top_right_text, filename, buyer_name):
    """解析发票数据"""
    # 查找发票代码和发票号码
    invoice_code_search = re.search(r"发票代码\s*[:：]?\s*(\d+)", first_page_text)
    invoice_number_search = re.search(r"发票号码\s*[:：]?\s*(\d+)", first_page_text)

    if not invoice_code_search and not invoice_number_search and top_right_text:
        invoice_code_search_in_top_right = re.search(r"发票代码[:：]\s{0,3}(\d+)", top_right_text)
        invoice_number_search_in_top_right = re.search(r"发票号码[:：]\s{0,3}(\d+)", top_right_text)

        invoice_code = invoice_code_search_in_top_right.group(1) if invoice_code_search_in_top_right else ""
        invoice_number = invoice_number_search_in_top_right.group(1) if invoice_number_search_in_top_right else ""
    else:
        invoice_code = invoice_code_search.group(1) if invoice_code_search else ""
        invoice_number = invoice_number_search.group(1) if invoice_number_search else ""

    # 查找金额
    # 打印原始文本，帮助调试
    print(f"\n=== 文件 {filename} 的文本内容 ===")
    print("First page text:")
    print(first_page_text)
    print("\nTop right text:")
    print(top_right_text)
    
    # 1. 先尝试查找"价税合计"后面的金额
    price_total_pattern = r"价税合计[¥￥·\s]*([\d,]+\.?\d*)"
    price_match = re.search(price_total_pattern, first_page_text)
    
    # 打印所有找到的金额
    all_amounts = re.findall(r"[￥¥·]\s{0,3}([\d,]+\.?\d*)", first_page_text)
    print("\n找到的所有金额:")
    for amount in all_amounts:
        print(f"- {amount}")
    
    if price_match:
        # 如果找到"价税合计"，直接使用这个金额
        amount_str = price_match.group(1).replace(',', '')
        total_amount = float(amount_str)
        print(f"\n使用价税合计金额: {total_amount}")
    else:
        # 2. 如果没找到，查找所有带货币符号的金额，取最大值
        amounts = re.findall(r"[￥¥·]\s{0,3}([\d,]+\.?\d*)", first_page_text)
        if amounts:
            # 移除逗号并转换为浮点数
            amounts_float = [float(amount.replace(',', '')) for amount in amounts]
            print("\n所有金额（浮点数）:")
            for amt in amounts_float:
                print(f"- {amt}")
            total_amount = max(amounts_float)
            print(f"\n使用最大金额: {total_amount}")
        else:
            print(f"\n未找到金额")
            total_amount = 0.0

    # 查找日期
    date_search = re.search(r"(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)", first_page_text + top_right_text)
    invoice_date = date_search.group(1) if date_search else "未找到发票日期"

    # 提取项目信息
    items = re.findall(r"\*(.*?)\*", full_text)
    simplified_item_names = ''.join(dict.fromkeys(''.join(re.findall(r'[\u4e00-\u9fa5]', item)) for item in items if item.strip()))

    # 提取购买方和销售方名称
    company_pattern = r"(?:名\s*称\s*[:：]\s*|^)(.+?(?:公司|事务所))"
    companies = re.findall(company_pattern, first_page_text, re.MULTILINE)
    
    if len(companies) >= 2:
        buyer_name = companies[0]
        seller_name = companies[1]
    else:
        buyer_name = buyer_name if buyer_name else "未找到购买方名称"
        seller_name = "未找到销售方名称"

    return {
        "filename": filename,
        "invoice_code": invoice_code,
        "invoice_number": invoice_number,
        "items": simplified_item_names,
        "total_amount": total_amount,
        "invoice_date": invoice_date,
        "buyer_name": buyer_name,
        "seller_name": seller_name
    }

def create_excel_report(data_list, temp_dir):
    """创建包含多个发票数据的Excel报告"""
    workbook = Workbook()
    sheet = workbook.active
    headers = ["文件名", "发票代码", "发票号码", "项目信息", "价税合计", "发票日期", "购买方名称", "销售方名称"]
    sheet.append(headers)
    
    # 设置列宽
    column_widths = [30, 15, 15, 40, 15, 20, 40, 40]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width
    
    # 添加所有发票数据
    for data in data_list:
        # 处理价税合计，确保是数字
        try:
            if isinstance(data["total_amount"], (int, float)):
                total_amount = float(data["total_amount"])
            else:
                amount_str = str(data["total_amount"])
                amount_str = ''.join(c for c in amount_str if c.isdigit() or c == '.')
                total_amount = float(amount_str)
        except (ValueError, TypeError):
            total_amount = 0.0
        
        # 创建行数据
        row_data = [
            data["filename"],
            data["invoice_code"],
            data["invoice_number"],
            data["items"],
            f"={total_amount}",  # 使用公式强制转换为数字
            data["invoice_date"],
            data["buyer_name"],
            data["seller_name"]
        ]
        
        # 添加行并设置格式
        sheet.append(row_data)
        current_row = sheet.max_row
        
        # 设置发票号码为文本
        invoice_cell = sheet.cell(current_row, 3)
        invoice_cell.data_type = TYPE_STRING
        invoice_cell.number_format = '@'
        
        # 设置价税合计为数字
        amount_cell = sheet.cell(current_row, 5)
        amount_cell.number_format = '#,##0.00'
    
    # 设置表头样式
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # 保存到临时目录
    excel_filename = f"Invoice_Data_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_path = os.path.join(temp_dir, excel_filename)
    workbook.save(excel_path)
    return excel_path, excel_filename

def process_invoice(file_storage, search_texts=None):
    """处理从网页上传的发票文件"""
    temp_dir = None
    pdf_path = None
    try:
        # 创建临时目录
        temp_dir = tempfile.mkdtemp()
        
        # 保存上传的文件
        filename = secure_filename(file_storage.filename)
        pdf_path = os.path.join(temp_dir, filename)
        file_storage.save(pdf_path)
        
        # 处理PDF文件
        with pdfplumber.open(pdf_path) as pdf:
            first_page_text, full_text, top_right_text = extract_invoice_data(pdf)
            
            # 如果提供了搜索文本，检查是否匹配
            if search_texts:
                matching_texts = [text for text in search_texts if text in full_text]
                if not matching_texts:
                    return {
                        'success': False,
                        'message': '未找到匹配的搜索文本'
                    }
                buyer_name = matching_texts[0]
            else:
                buyer_name = ""
            
            # 解析发票数据
            result = parse_data(first_page_text, full_text, top_right_text, filename, buyer_name)
        
        # 生成Excel报告
        excel_path, excel_filename = create_excel_report(result, temp_dir)
        
        # 只删除PDF文件，保留Excel文件
        if pdf_path and os.path.exists(pdf_path):
            os.remove(pdf_path)
        
        return {
            'success': True,
            'data': result,
            'excel_filename': excel_filename,
            'excel_path': excel_path,
            'temp_dir': temp_dir,  # 添加临时目录路径
            'message': '发票处理成功'
        }
        
    except Exception as e:
        # 清理文件
        try:
            if pdf_path and os.path.exists(pdf_path):
                os.remove(pdf_path)
            if temp_dir and os.path.exists(temp_dir):
                # 尝试删除目录中的所有文件
                for file in os.listdir(temp_dir):
                    os.remove(os.path.join(temp_dir, file))
                os.rmdir(temp_dir)
        except Exception as cleanup_error:
            pass
            
        return {
            'success': False,
            'error': str(e),
            'message': '发票处理失败'
        }

def process_multiple_invoices(files, search_texts=None):
    """处理多个发票文件"""
    processed_invoices = set()  # 用于存储已处理的发票代码和号码组合
    results = []
    duplicate_count = 0
    processed_filenames = []    
    skipped_filenames = []      
    duplicate_filenames = []    
    
    try:
        for file in files:
            if len(file.read()) > MAX_FILE_SIZE:
                return {
                    'success': False,
                    'message': f'文件 {file.filename} 超过大小限制'
                }
            file.seek(0)  # 重置文件指针
            
            try:
                # 直接从内存处理PDF，不保存到磁盘
                pdf_bytes = file.read()
                pdf_file = io.BytesIO(pdf_bytes)
                
                with pdfplumber.open(pdf_file) as pdf:
                    first_page_text, full_text, top_right_text = extract_invoice_data(pdf)
                    
                    # 如果提供了搜索文本，检查是否匹配
                    if search_texts:
                        matching_texts = [text for text in search_texts if text in full_text]
                        if not matching_texts:
                            skipped_filenames.append(file.filename)  # 记录不匹配的文件
                            continue
                    
                    # 解析发票数据，传入原始文件名
                    result = parse_data(first_page_text, full_text, top_right_text, file.filename, "")
                    
                    # 检查重复发票
                    invoice_key = (result['invoice_code'], result['invoice_number'])
                    if invoice_key in processed_invoices:
                        duplicate_count += 1
                        duplicate_filenames.append(file.filename)  # 记录重复文件名
                        continue
                    
                    processed_invoices.add(invoice_key)
                    results.append(result)
                    processed_filenames.append(file.filename)
                
                # 删除处理完的PDF文件
                os.remove(pdf_path)
                
            except Exception as e:
                skipped_filenames.append(file.filename)  # 记录处理失败的文件
                continue
        
        if not results:
            return {
                'success': False,
                'message': '没有找到有效的发票数据'
            }
        
        # 生成Excel报表到内存
        excel_buffer = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        headers = ["文件名", "发票代码", "发票号码", "项目信息", "价税合计", "发票日期", "购买方名称", "销售方名称"]
        sheet.append(headers)
        
        # 设置列宽
        column_widths = [30, 15, 15, 40, 15, 20, 40, 40]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = width
        
        # 添加所有发票数据
        for data in results:
            # 处理价税合计，确保是数字
            try:
                if isinstance(data["total_amount"], (int, float)):
                    total_amount = float(data["total_amount"])
                else:
                    amount_str = str(data["total_amount"])
                    amount_str = ''.join(c for c in amount_str if c.isdigit() or c == '.')
                    total_amount = float(amount_str)
            except (ValueError, TypeError):
                total_amount = 0.0
            
            # 创建行数据
            row_data = [
                data["filename"],
                data["invoice_code"],
                data["invoice_number"],
                data["items"],
                f"={total_amount}",  # 使用公式强制转换为数字
                data["invoice_date"],
                data["buyer_name"],
                data["seller_name"]
            ]
            
            # 添加行并设置格式
            sheet.append(row_data)
            current_row = sheet.max_row
            
            # 设置发票号码为文本
            invoice_cell = sheet.cell(current_row, 3)
            invoice_cell.data_type = TYPE_STRING
            invoice_cell.number_format = '@'
            
            # 设置价税合计为数字
            amount_cell = sheet.cell(current_row, 5)
            amount_cell.number_format = '#,##0.00'
        
        # 设置表头样式
        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        workbook.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        
        return {
            'success': True,
            'total_processed': len(results),
            'duplicate_count': duplicate_count,
            'excel_content': excel_content,
            'excel_filename': f'Invoice_Data_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx',
            'processed_filenames': processed_filenames,
            'skipped_filenames': skipped_filenames,
            'duplicate_filenames': duplicate_filenames
        }
        
    except Exception as e:
        print(f"处理失败: {str(e)}")
        return {
            'success': False,
            'message': f'处理失败: {str(e)}'
        }
